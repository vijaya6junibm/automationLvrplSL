import csv
import pathlib
import sys
from pathlib import Path
import os
from openpyxl import Workbook, load_workbook
from contextlib import closing
import string
import json
import xmltodict
from xml.sax.saxutils import escape

def get_data_file_path():
    """Gets the file path of the data directory
    Args:
    cwd - gets the file path of the current working directory
    parent - the folder name of the parent directory
    """

    cwd = os.getcwd()
    parent = os.path.basename(os.path.dirname(cwd))
    print('cwd')
    print(cwd)
    print('parent')
    print(parent)


def make_excel_file(file_name):
    with closing(Workbook()) as wb:
        wb.save(file_name)

def append_excel_file1(file_name1,order_no,test_name,timeStamp,inputXML,outputXML):
    with open(file_name1, 'a', newline='') as csvfile:
        writer = csv.writer(csvfile)  # getting a csv.writer object
        csvdata = (test_name,timeStamp,order_no,inputXML,outputXML)
        #csvdata1 = ('TestCaseName', test_name)
        writer.writerow(csvdata)  # appending a line to the end file. csvdata is a tuple
        #writer.writerow(csvdata1)


def where() -> pathlib.Path:
    """Return the path the directory of this file
    """
    this_dir = pathlib.Path(__file__).parent
    script_directory = os.path.dirname(os.path.abspath(sys.argv[0]))
    return this_dir

def where1() -> str:
    """Return the path the directory of this file
    """
    #this_dir = pathlib.Path(__file__).parent
    script_directory = os.path.dirname(os.path.abspath(sys.argv[0]))
    return script_directory

def ord_mig(file_name1,order_no,ord_header_key):
    with open(file_name1, 'a', newline='') as csvfile:
        writer = csv.writer(csvfile)  # getting a csv.writer object
        csvdata = (order_no,ord_header_key)
        #csvdata1 = ('TestCaseName', test_name)
        writer.writerow(csvdata)  # appending a line to the end file. csvdata is a tuple
        #writer.writerow(csvdata1)

def append_excel_file2(file_name1,order_no,order_header_key):
    with open(file_name1, 'a', newline='') as csvfile:
        writer = csv.writer(csvfile)  # getting a csv.writer object
        csvdata = (order_no,order_header_key)
        #csvdata1 = ('TestCaseName', test_name)
        writer.writerow(csvdata)  # appending a line to the end file. csvdata is a tuple
        #writer.writerow(csvdata1)

def create_order_input_file(item_id):
    dirpath = 'C:/EPL-Liverpool-SL/EPLOpenOrdersMigration'
    path1 = 'ItemDynamic'
    with open(dirpath+'/'+path1+'/Input/CreateOrderInput.xml') as fd:
        doc = fd.read()
        doc = doc.replace('$OrderNo', item_id)
        doc = doc.replace('$ItemID', item_id)
    with open(dirpath + '/' + path1 + '/Input/inputWithOrderId.xml', 'w', newline='') as fd1:
        fd1.write(doc)
        print(doc)
    return doc


def schedule_input_file1_del(order_no,Order_Header_Key,folder_path):
    folder_path1 = "\\".join(folder_path.split("\\")[0:-1])
    tpath = folder_path1 + '/Input/scheduleOrder.xml'
    inner_template = string.Template(
        '    <ScheduleOrder DocumentType="${DocumentType}" EnterpriseCode="${EnterpriseCode}" OrderHeaderKey="${OrderHeaderKey}" OrderNo="${OrderNo}"></ScheduleOrder>')
    outer_template = string.Template("""
    ${document_list}
     """)
    # <ScheduleOrder  DocumentType="0001" EnterpriseCode="Liverpool" OrderHeaderKey="20241203090251475620" OrderNo="0904202401"/>
    data = [('0001', 'Liverpool', Order_Header_Key, order_no)]
    inner_contents = [inner_template.substitute(DocumentType=DocumentType, EnterpriseCode=EnterpriseCode,
                                                OrderHeaderKey=OrderHeaderKey, OrderNo=OrderNo) for
                      (DocumentType, EnterpriseCode, OrderHeaderKey, OrderNo) in data]
    result = outer_template.substitute(document_list='\n'.join(inner_contents))
    print(result)
    with open(folder_path1, 'w', newline='') as fd:
        fd.write(result)
        return result

def release_input_file1_del(order_no,Order_Header_Key):
    inner_template = string.Template(
        '    <ReleaseOrder DocumentType="${DocumentType}" EnterpriseCode="${EnterpriseCode}" OrderHeaderKey="${OrderHeaderKey}" OrderNo="${OrderNo}"></ReleaseOrder>')
    outer_template = string.Template("""
    ${document_list}
     """)
    # <ScheduleOrder  DocumentType="0001" EnterpriseCode="Liverpool" OrderHeaderKey="20241203090251475620" OrderNo="0904202401"/>
    data = [('0001', 'Liverpool', Order_Header_Key, order_no)]
    inner_contents = [inner_template.substitute(DocumentType=DocumentType, EnterpriseCode=EnterpriseCode,
                                                OrderHeaderKey=OrderHeaderKey, OrderNo=OrderNo) for
                      (DocumentType, EnterpriseCode, OrderHeaderKey, OrderNo) in data]
    result = outer_template.substitute(document_list='\n'.join(inner_contents))
    print(result)
    with open('input/ReleaseOrder.xml', 'w', newline='') as fd:
        fd.write(result)
        return result

def somsns_input_file1_del(order_no, ItemID):
        inner_template = string.Template(
            '    <OrderStatusChange OrderNo="${OrderNo}"><OrderLines><OrderLine ItemID=${ItemID} BaseDropStatus="CONF"></OrderLine></OrderLines></OrderStatusChange>')
        outer_template = string.Template("""
        ${document_list}
         """)
        #<OrderStatusChange OrderNo="OtestItem9a">
	#<OrderLines>
		#<OrderLine ItemID="testItem9a" BaseDropStatus="CONF"/>
	#</OrderLines>
#</OrderStatusChange>
        data = [('0001', 'Liverpool', ItemID, order_no)]
        inner_contents = [inner_template.substitute(DocumentType=DocumentType, EnterpriseCode=EnterpriseCode,
                                                    ItemID=ItemID, OrderNo=OrderNo) for
                          (DocumentType, EnterpriseCode, ItemID, OrderNo) in data]
        result = outer_template.substitute(document_list='\n'.join(inner_contents))
        print(result)
        with open('input/somsns.xml', 'w', newline='') as fd:
            fd.write(result)
            return result

def getOrderDetails_input_file1_del(order_no,Order_Header_Key):
    inner_template = string.Template(
        '    <Order DocumentType="${DocumentType}" EnterpriseCode="${EnterpriseCode}" OrderHeaderKey="${OrderHeaderKey}" OrderNo="${OrderNo}"></Order>')
    outer_template = string.Template("""
    ${document_list}
     """)
    # <ScheduleOrder  DocumentType="0001" EnterpriseCode="Liverpool" OrderHeaderKey="20241203090251475620" OrderNo="0904202401"/>
    data = [('0001', 'Liverpool', Order_Header_Key, order_no)]
    inner_contents = [inner_template.substitute(DocumentType=DocumentType, EnterpriseCode=EnterpriseCode,
                                                OrderHeaderKey=OrderHeaderKey, OrderNo=OrderNo) for
                      (DocumentType, EnterpriseCode, OrderHeaderKey, OrderNo) in data]
    result = outer_template.substitute(document_list='\n'.join(inner_contents))
    print(result)
    with open('input/getOrderDetails.xml', 'w', newline='') as fd:
        fd.write(result)
        return result

def write_output(output1,name):
    upload_folder = os.path.abspath("Output")
    with open(upload_folder+'/'+name+'.json', 'w', newline='') as fd:
        fd.write(output1)

def append_excel_file2(scheduleOrder):
    with open(scheduleOrder, 'a', newline='') as fd:
        writer = csv.writer(fd)
        doc = fd.read()
        return doc

def add_value(file_name, cell_cords, value):
    with closing(load_workbook(filename=file_name)) as wb:
        ws = wb.active
        ws[cell_cords] = value
        wb.save(file_name)